#! /usr/bin/env python3

import openpyxl, os, sys

##print msg with color
def print_msg(msg, color=-1):
    if color == 0:  #0 green
        print('\033[1;32;40m', end='')
        print(msg, end='')
        print('\033[0m')
    elif color == 1:  #1 red
        print('\033[1;31;40m', end='')
        print(msg, end='')
        print('\033[0m')
    elif color == 2:  #2 blue
        print('\033[1;34;40m', end='')
        print(msg, end='')
        print('\033[0m')
    else:
        print(msg)

##parse excel file 
def get_excel_string(excelFile, langIndex):
    workbook = openpyxl.load_workbook(excelFile)
    sheet = workbook.get_active_sheet()
    rowsMax = 1048576
    old_strArr = []
    new_strArr = []
    for curRow in range(1, rowsMax):
        columnAValue = sheet.cell(row = curRow, column=1).value
        old_str = str(columnAValue) + '_'+ langIndex +'.0=' #"_3.0="
        if columnAValue == None:
            break;
        else:
            colBValue = sheet.cell(row = curRow, column = 2).value
            new_str = old_str + str(colBValue)
            old_strArr.append(old_str)
            new_strArr.append(new_str)

    return (old_strArr, new_strArr)

def file_str_replace(excelFile, estFile, langIndex):
    try:
        file = open(estFile)
        fileStr = file.read()
        fileList = fileStr.split("\n")

        (old_str_arr, new_str_arr) = get_excel_string(excelFile, langIndex)

        for i in range(0, len(old_str_arr)):
            for strTemp in fileList:
                if old_str_arr[i] in strTemp:
                    index = fileList.index(strTemp)
                    fileList[index] = new_str_arr[i]

        fileStr = "\n".join(fileList)
        file.close()
        file = open(estFile, 'w')
        file.write(fileStr)
    except :
        #print('\033[1;31;40m', end='') 
        #print(estFile + " Transfer Failed!!!")
        #print('\033[0m')
        print_msg(estFile + " Transfer Failed!!!", 1)
        return

    print_msg(estFile + " Conversion Success!", 0)
    #print('\033[1;32;40m', end='') 
    #print(estFile + " Conversion Success!")
    #print('\033[0m')

##get input param 
def isIntNumber(s):
    try:
        num = int(s)
        if (num > 0):
            return True
        else:
            return False
    except ValueError:
        pass

    try:
        import unicodedata
        unicodedata.numeric(s)
        return True
    except (TypeError, ValueError):
        pass

    return False

def getInputPara():
    flag = 0
    for i in range(1, len(sys.argv)):
        if (".xls" in sys.argv[i] or ".xlsx" in sys.argv[i]):
            flag = flag + 1
            excelFile = sys.argv[i]
        if (".est" in sys.argv[i]):
            flag = flag + 1
            estFile = sys.argv[i]
            print_msg("Est File:" + sys.argv[i])
            #print("Est File:" + sys.argv[i])
            #print("Excel File:" + sys.argv[i])
        if (isIntNumber(sys.argv[i])):
            flag = flag + 1
            languageIndex = sys.argv[i]
            print_msg("Language index:" + sys.argv[i])
            #print("Language index:" + sys.argv[i])
    if (flag == 3):
        print_msg("The command has received and is translating. Please wait...", 2)
        #print("The command has received and is translating. Please wait...")
        return [excelFile, estFile, languageIndex]
    else:
        print_msg("The parameters you entered are incorrect!!", 1)
        #print("The parameters you entered are incorrect!!")
        return ['', '', '']


ExcelFile = ''
EstFile = ''
LangIndex = ''
paramTable = [ExcelFile, EstFile, LangIndex]

paramTable = getInputPara()
#print(paramTable)

pwd = os.getcwd()
ExcelFile = pwd + '/' + paramTable[0]
EstFile = pwd + '/' + paramTable[1]
LangIndex = paramTable[2]

#print(ExcelFile, EstFile, LangIndex)

file_str_replace(ExcelFile, EstFile, LangIndex)
